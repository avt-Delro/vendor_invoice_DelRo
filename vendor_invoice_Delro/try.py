# # Your Aspose credentials
# client_id = "a26aee91-9272-461e-9439-09001f6e99cc"
# client_secret = "1722b75a2d2bc5f0091e3e4bdd7c08af"

# # Get access token
# token_url = "https://api.aspose.cloud/connect/token"
# payload = {
#     "grant_type": "client_credentials",
#     "client_id": client_id,
#     "client_secret": client_secret
# }
# token_response = requests.post(token_url, data=payload)
# access_token = token_response.json()["access_token"]

# # Download the file directly from Aspose Cloud storage
# file_path = "invoice22_invoices.xlsx"
# download_url = f"https://api.aspose.cloud/v3.0/storage/file/{file_path}"
# headers = {"Authorization": f"Bearer {access_token}"}
# file_response = requests.get(download_url, headers=headers)
# output_path = r"C:\Users\VJDELROSARIO\OneDrive - ATV INC\Desktop\Automation\vendor_invoice_Delro\files_downloaded"

# with open(f"{output_path}\{file_path}", "wb") as f:
#     f.write(file_response.content)


# def online_conversion():
#     config = Configuration()
#     config.app_sid = "a26aee91-9272-461e-9439-09001f6e99cc"
#     config.app_key = "1722b75a2d2bc5f0091e3e4bdd7c08af"
#     api_client = ApiClient(app_sid=config.app_sid, app_key=config.app_key)
#     pdf_api = PdfApi(api_client)

#     full_path = f"{folder_path_upload}\{list_of_shops[0]}"

#     for filename in os.listdir(full_path):
#         if filename.endswith(".pdf"):
#             local_path = os.path.join(full_path, filename)
#             remote_path = f"{remote_folder_pdf}/{filename}"
#             pdf_api.upload_file(remote_path, local_path)
#             print(f"Uploaded {filename} to {remote_path}")

#     files_list_response = pdf_api.get_files_list(remote_folder_pdf)

#     for file_in in files_list_response.value:
#         filenm = file_in.name
#         if filenm.lower().endswith(".pdf"):
#             remote_pdf = f"{remote_folder_pdf}\{filenm}"
#             remote_excel = f"{remote_folder_excel}/{filenm.replace(".pdf", ".xlsx")}"

#             pdf_api.put_pdf_in_storage_to_xlsx(remote_pdf, remote_excel)

from openpyxl import load_workbook

def mergeexcel():
    # Load the workbook and access Sheet 2
    ex_file = r"C:\Users\VJDELROSARIO\OneDrive - ATV INC\Desktop\Automation\vendor_invoice_Delro\test\invoice22_invoices.xlsx"
    wb = load_workbook(ex_file)
    sheet1 = wb.worksheets[0]  # Sheet index starts at 0
    sheet2 = wb.worksheets[1]  # Sheet index starts at 0

    for sheet in wb.worksheets:
        # Copy merged ranges to avoid modifying while iterating
        merged_ranges = list(sheet.merged_cells.ranges)

        for merged_range in merged_ranges:
            # Get the value from the top-left cell of the merged range
            retained_value = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value

            # Unmerge the range
            sheet.unmerge_cells(str(merged_range))

            # Write the value only to the rightmost cell in the merged range
            rightmost_cell = sheet.cell(row=merged_range.min_row, column=merged_range.max_col)
            rightmost_cell.value = retained_value


    target_sheet = wb.worksheets[0]  # Assuming you want to consolidate into the first sheet

    for sheet in wb.worksheets[1:]:
        # Find the current last row in the target sheet
        last_row = target_sheet.max_row

        # Copy data from the current sheet to the target sheet
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=last_row):
            for col_idx, value in enumerate(row, start=1):
                target_sheet.cell(row=row_idx, column=col_idx, value=value)

    wb.save(ex_file)
    print("Unmerged all merged cells in the workbook.")

def findthis(search_val):
    search_text_inv = False
    collected_inv = ""
    for row in sheet.iter_rows(min_col=5, max_col=6):
        cell = row[0]
        next_cell = row[1]
        value = str(cell.value) if cell.value else ""
        nextval = str(next_cell.value) if next_cell.value else ""

        if not search_text_inv:
            if value == search_val:
                search_text_inv = True
                collected_inv = nextval

    print(f"{search_val} : {collected_inv}")

ex_file = r"C:\Users\VJDELROSARIO\OneDrive - ATV INC\Desktop\Automation\vendor_invoice_Delro\test\invoice22_invoices.xlsx"
wb = load_workbook(ex_file)
sheet = wb.active

search_text_found = False
collected_text_bill_to = ""

for row in sheet.iter_rows(min_col=1, max_col=1):
    cell = row[0]
    value = str(cell.value) if cell.value else ""

    if not search_text_found:
        if value == "BILL TO":
            search_text_found = True
            print(f"Found at {cell.coordinate}")
    else:
        print(f"{cell.coordinate} :: {cell.value}")
        if value == "SKU":
            print(f"found {cell.coordinate}")
            break
        collected_text_bill_to += value + " "

print(collected_text_bill_to)

search_text_found1 = False
collected_text1_ship_to = ""

for row in sheet.iter_rows(min_col=2, max_col=2):
    cell = row[0]
    value = str(cell.value) if cell.value else ""

    if not search_text_found1:
        if value == "SHIP TO":
            search_text_found1 = True
            print(f"Found at {cell.coordinate}")
    else:
        print(f"{cell.coordinate} :: {cell.value}")
        if value == "DESCRIPTION":
            print(f"found {cell.coordinate}")
            break
        collected_text1_ship_to += value + " "

print(collected_text1_ship_to)

trylist = ["INVOICE", "DATE", "TERMS", "DUE DATE"]

for i in trylist:
    findthis(i)







    
        
    






