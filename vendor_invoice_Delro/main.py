import os
import pandas as pd
from openpyxl import Workbook
from asposepdfcloud import PdfApi, ApiClient
from asposepdfcloud.configuration import Configuration
import requests
import aspose.pdf as ap
from aspose.pdf import SaveFormat
import pdfplumber
from openpyxl import load_workbook



list_of_shops = ["A2Z"]

def findthis(search_val, wb, wb_sheet):
    excel_file = load_workbook(wb)
    wb_sheet = excel_file.active
    search_text_inv = False
    collected_inv = ""
    for row in wb_sheet.iter_rows(min_col=5, max_col=6):
        cell = row[0]
        next_cell = row[1]
        value = str(cell.value) if cell.value else ""
        nextval = str(next_cell.value) if next_cell.value else ""

        if not search_text_inv:
            if value == search_val:
                search_text_inv = True
                collected_inv = nextval

    print(f"{search_val} : {collected_inv}")



def excel_mapping(shop_name):
    folder_path = r"C:\Users\VJDELROSARIO\OneDrive - ATV INC\Desktop\Automation\vendor_invoice_Delro\files_downloaded"
    if os.path.exists(folder_path) == True:
        shop_folder_path = os.path.join(folder_path, shop_name)
        print(shop_folder_path)
        for file in os.listdir(shop_folder_path):
            if file.endswith(".xlsx"):
                file_excel = load_workbook(os.path.join(shop_folder_path, file))
                sheet1 = file_excel.active

                search_text_found = False
                collected_text_bill_to = ""

                for row in sheet1.iter_rows(min_col=1, max_col=1):
                    cell = row[0]
                    value = str(cell.value) if cell.value else ""

                    if not search_text_found:
                        if value == "BILL TO":
                            search_text_found = True
                            print(f"Found at {cell.coordinate}")
                    else:
                        print(f"{cell.coordinate} :: {cell.value}")
                        if value == "SKU":
                            print(f"found {cell.coordinate}")
                            break
                        collected_text_bill_to += value + " "

                print(collected_text_bill_to)

                search_text_found1 = False
                collected_text1_ship_to = ""

                for row in sheet1.iter_rows(min_col=2, max_col=2):
                    cell = row[0]
                    value = str(cell.value) if cell.value else ""

                    if not search_text_found1:
                        if value == "SHIP TO":
                            search_text_found1 = True
                            print(f"Found at {cell.coordinate}")
                    else:
                        print(f"{cell.coordinate} :: {cell.value}")
                        if value == "DESCRIPTION":
                            print(f"found {cell.coordinate}")
                            break
                        collected_text1_ship_to += value + " "

                print(collected_text1_ship_to)

                trylist = ["INVOICE", "DATE", "TERMS", "DUE DATE"]
                for i in trylist:
                    findthis(i, os.path.join(shop_folder_path, file), sheet1)


            else:
                print(f"{file} files are not supported.")
    else:
        print(f"Folder for does not exist. Skipping...")
        return

#This function checks if directory is present, creates if none
def createdirect(shop_name):
    folder_to_be_uploaded = r"C:\Users\VJDELROSARIO\OneDrive - ATV INC\Desktop\Automation\vendor_invoice_Delro\files_to_upload"
    shop_folder_to_be_uploaded = os.path.join(folder_to_be_uploaded, shop_name)

    if os.path.exists(shop_folder_to_be_uploaded) == False:
        os.makedirs(shop_folder_to_be_uploaded)
    else:
        print(f"Folder directory exists, upload file here {shop_folder_to_be_uploaded}")

#This function checks if the invoice is a multi-page document or not
def is_multipage_pdf(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        if len(pdf.pages) > 1:
            return True
        else:
            return False     

#This function converts the pdf to excel, in order to be mapped   
def convert_pdf_to_excel(shop_name):
    folder_path_upload = r"C:\Users\VJDELROSARIO\OneDrive - ATV INC\Desktop\Automation\vendor_invoice_Delro\files_to_upload"
    output_path = r"C:\Users\VJDELROSARIO\OneDrive - ATV INC\Desktop\Automation\vendor_invoice_Delro\files_downloaded"

    shop_folder = os.path.join(folder_path_upload, shop_name)
    shop_upload_path = os.path.join(output_path, shop_name)
    
    if os.path.exists(shop_upload_path) == False:
        os.makedirs(shop_upload_path)

    for filename in os.listdir(shop_folder):
        if filename.endswith(".pdf"):
            is_multi = is_multipage_pdf(os.path.join(shop_folder, filename))
            if is_multi:
                print(f"{filename} is a multi-page PDF.")
                pdf_path = ap.Document(os.path.join(shop_folder, filename))
                pdf_path.save(os.path.join(shop_upload_path, filename.replace(".pdf", ".xlsx")), ap.SaveFormat.EXCEL)
                print(f"Converted {filename} to Excel format.")
                #This is needed in order to join the sheets together
                the_excel_file = os.path.join(shop_upload_path, filename.replace(".pdf", ".xlsx"))
                wb = load_workbook(the_excel_file)
                sheet_names = wb.worksheets
                for sheet in sheet_names:
                    merged_ranges = list(sheet.merged_cells.ranges)

                    for merged_range in merged_ranges:
                    # Get the value from the top-left cell of the merged range
                        retained_value = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value

                        # Unmerge the range
                        sheet.unmerge_cells(str(merged_range))

                        # Write the value only to the rightmost cell in the merged range
                        rightmost_cell = sheet.cell(row=merged_range.min_row, column=merged_range.max_col)
                        rightmost_cell.value = retained_value
                
                target_sheet = wb.worksheets[0] 
                for sheet in wb.worksheets[1:]:
                    # Find the current last row in the target sheet
                    last_row = target_sheet.max_row

                    # Copy data from the current sheet to the target sheet
                    #Starts at row 2 after the copyright, of Aspose
                    for row_idx, row in enumerate(sheet.iter_rows(values_only=True, min_row=2), start=last_row):
                        for col_idx, value in enumerate(row, start=1):
                            target_sheet.cell(row=row_idx, column=col_idx, value=value)
                
                wb.save(the_excel_file)
                print(f"File {filename} a multi-sheet xlsx is combined as one")
                   
            if not is_multi:
                print(f"{filename} is a single-page PDF.")
                pdf_path = ap.Document(os.path.join(shop_folder, filename))
                pdf_path.save(os.path.join(shop_upload_path, filename.replace(".pdf", ".xlsx")), ap.SaveFormat.EXCEL)

                the_excel_file = os.path.join(shop_upload_path, filename.replace(".pdf", ".xlsx"))
                wb = load_workbook(the_excel_file)
                sheet_names = wb["Sheet1"]
                for merged_range in list(sheet_names.merged_cells.ranges):
                    # Get the value from the top-left cell of the merged range
                    retained_value = sheet_names.cell(row=merged_range.min_row, column=merged_range.min_col).value

                    # Only unmerge if the range is still merged
                    if merged_range in sheet_names.merged_cells.ranges:
                        sheet_names.unmerge_cells(str(merged_range))

                        # Write the value to the rightmost cell
                        rightmost_cell = sheet_names.cell(row=merged_range.min_row, column=merged_range.max_col)
                        rightmost_cell.value = retained_value
                print(f"Converted {filename} to Excel format.")

def main():
    for shop in list_of_shops:
        createdirect(shop)
        convert_pdf_to_excel(shop)
    # excel_mapping(list_of_shops[0])

main()

